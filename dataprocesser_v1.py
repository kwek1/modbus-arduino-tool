from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    ScatterChart,
    Reference,
    Series,
    marker
)
from PyQt5.QtCore import QObject, pyqtSignal


class dataProcesser(QObject):
    finished = pyqtSignal()
    progress = pyqtSignal(str)

    def __init__(self, data, file_path):
        super().__init__()
        self.arduino_data = data['arduino']
        self.modbus_data = data['modbus']
        self.file_path = file_path
        self.suffix = data['run_num']

    def __fix(self, arr):
        def helper(x):
            if x > 65500:
                x -= 65535
            return x/10
        return (arr[0], helper(arr[1]), helper(arr[2]))

    def run(self):
        wb = Workbook()

        # log flow readings
        ws = wb.create_sheet('flow data', 1)
        ws["A1"] = 'timestamp'
        ws['B1'] = 'reading'
        for i in range(len(self.arduino_data)):
            ws[f"A{i+2}"] = self.arduino_data[i][0]
            ws[f"B{i+2}"] = self.arduino_data[i][1]
        chart = LineChart()
        chart_data = Reference(ws, min_col=2, min_row=2,
                               max_row=len(self.arduino_data))  # does not plot the last entry
        chart.add_data(chart_data)
        ws.add_chart(chart, "C2")
        #self.progress.emit("Created flow graph")

        # log pressure readings
        ws = wb.create_sheet('pressure data', 2)
        ws["A1"] = 'timestamp'
        ws['B1'] = 'Inlet reading'
        ws['C1'] = 'Outlet reading'
        self.modbus_data = [i for i in map(self.__fix, self.modbus_data)]
        for i in range(len(self.modbus_data)):
            ws[f"A{i+2}"] = self.modbus_data[i][0]
            ws[f"B{i+2}"] = self.modbus_data[i][1]
            ws[f"C{i+2}"] = self.modbus_data[i][2]
        chart = LineChart()
        chart_data = Reference(ws, min_col=2, max_col=3,
                               min_row=2, max_row=len(self.modbus_data))  # does not plot the last entry
        chart.add_data(chart_data)
        ws.add_chart(chart, "D2")
        #self.progress.emit("Created pressure graph")

        # collate data by timestamp
        # basically, for each reading from the arduino, we look through the readings from the modbus
        # and match the one with the timestamp nearest to it

        ws = wb["Sheet"]
        ws.title = "Collated data"
        ws["A1"] = 'Flow timestamp'
        ws['B1'] = 'Flow reading'
        ws['C1'] = 'Inlet pressure reading'
        ws['D1'] = 'Outlet pressure reading'
        ws["E1"] = 'Pressure timestamp'

        def find_nearest_idx(arr, target):
            arr = [abs(i - target) for i in arr]
            return arr.index(min(arr))

        time_list = [i[0] for i in self.modbus_data]
        idx_list = [find_nearest_idx(time_list, i[0])
                    for i in self.arduino_data]

        row_num = 2
        for idx in idx_list:
            modbus_entry = [self.modbus_data[idx][1],
                            self.modbus_data[idx][2],
                            self.modbus_data[idx][0]]
            row_entry = [i for i in self.arduino_data[row_num - 2]]
            row_entry.extend(modbus_entry)
            for i in range(1, 6):
                ws.cell(row=row_num, column=i, value=row_entry[i-1])
            row_num += 1

        chart = ScatterChart()  # x-axis is outlet pressure, y-axis is flow
        x_data = Reference(ws, min_col=4, min_row=2,
                           max_row=len(self.arduino_data))
        y_data = Reference(ws, min_col=2, min_row=2,
                           max_row=len(self.arduino_data))
        chart.title = "flow rate (l/min) vs p out (psi)"
        chart.y_axis.title = 'L/min'
        chart.x_axis.title = 'PSI'
        series = Series(y_data, x_data)
        series.marker = marker.Marker('x')
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
        ws.add_chart(chart, "E3")
        #self.progress.emit("Created overall graph")

        wb.save(f"{self.file_path}/Flow_pressure_data_run_{self.suffix}.xlsx")
        self.progress.emit("data saved")
        self.finished.emit()
